import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_inventory_per_supplier = {}
products_under_10 = {}

print(product_list.max_row)

for row in range(2, product_list.max_row+1):
    # getting values from a row and column in spreadsheet
    supplier_name = product_list.cell(row, 4).value
    inventory = product_list.cell(row, 2).value
    price = product_list.cell(row, 3).value
    product_num = product_list.cell(row, 1).value
    inventory_price = product_list.cell(row, 5)

    # calculating number of products per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier[supplier_name] + 1
    else:
        print("adding a new supplier")
        product_per_supplier[supplier_name] = 1

    # calculating total value of inventory for each supplier as price * inventory
    if supplier_name in total_inventory_per_supplier:
        current_value = total_inventory_per_supplier.get(supplier_name)
        total_inventory_per_supplier[supplier_name] = int(current_value) + int(inventory * price)
    else:
        total_inventory_per_supplier[supplier_name] = int(inventory * price)


    # getting all products with inventory less than 10
    if inventory < 10:
        products_under_10[int(product_num)] = int(inventory)

    # adding a value to a new column 5
    inventory_price.value = inventory * price


print(product_per_supplier)
print(total_inventory_per_supplier)
print(products_under_10)

inv_file.save("inventory_updated.xlsx")